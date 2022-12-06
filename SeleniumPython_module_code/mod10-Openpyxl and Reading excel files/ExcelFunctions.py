'''
Created on 05-Mar-2020
@author: Jaspreet
'''
import openpyxl 
class ReadWrite:
    wb=None
    filePath= None
    
    def __init__(self, path):
        global wb
        global filePath
        filePath = path
        wb = openpyxl.load_workbook(path)
        
    def readby_ColumnIndex(self, sheetname, rowIndex, colIndex):
        sheet = wb[sheetname]
        return sheet.cell(rowIndex,colIndex).value
    
    def readby_ColumnName(self,sheetname, rowIndex,colName):
        sheet = wb[sheetname]
        colIndex=1
        while(sheet.cell(row=1,column=colIndex).value!=''):
            if(colName==sheet.cell(row=1,column=colIndex).value):
                break
            colIndex = colIndex+1
        return sheet.cell(rowIndex, colIndex).value
    
    def writeby_ColumIndex(self,sheetname, rowIndex, colIndex,value):
        sheet = wb[sheetname]
        sheet.cell(row=rowIndex,column=colIndex).value = value
        wb.save(filePath)
        
    def writeby_ColumnName(self,sheetname, rowIndex, colName,value):
        sheet = wb[sheetname]
        colIndex=1
        while(sheet.cell(row=1,column=colIndex).value!=''):
            if(colName==sheet.cell(row=1,column=colIndex).value):
                break
            colIndex = colIndex+1
            sheet.cell(row=rowIndex,column=colIndex).value=value
            wb.save(filePath)
            
    def getRowCount(self,sheetname):
        sheet = wb[sheetname]
        return sheet.max_row
    
    def getColumnCount(self,sheetname):
        sheet = wb[sheetname]
        return sheet.max_column